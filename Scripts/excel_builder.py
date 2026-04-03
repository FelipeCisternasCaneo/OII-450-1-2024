import os
import pandas as pd
import re

from openpyxl.styles import Alignment, PatternFill
from openpyxl.formatting.rule import ColorScaleRule 
from openpyxl.drawing.image import Image

import matplotlib.pyplot as plt
import seaborn as sns
from scipy import stats

# --- CONFIGURACIÓN ---
RESUMEN_DIR = 'Resultados/resumen/KP/'
OUTPUT_EXCEL = 'Reporte_Final_Visual_KP_v2.xlsx'
PARAM_POBLACION = 10
PARAM_ITERACIONES = 100
PARAM_PROBLEMA = 'KP'
# --------------------

# --- FUNCIONES DE ESTILO (MODIFICADAS) ---
def aplicar_estilos_basicos(worksheet):
    center_alignment = Alignment(horizontal='center', vertical='center')
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = center_alignment
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells) + 2
        worksheet.column_dimensions[column_cells[0].column_letter].width = length

def aplicar_estilos_avanzados_resumen(worksheet, df_estadisticas, start_row_stats, end_row_stats):
    fill_gvp = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_std = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    formato_numero = '#,##0.00'
    for row_idx in range(start_row_stats, end_row_stats + 1):
        celda_binarizacion = worksheet.cell(row=row_idx, column=1)
        if celda_binarizacion.value and 'GVP' in celda_binarizacion.value:
            for cell in worksheet[row_idx]: cell.fill = fill_gvp
        elif celda_binarizacion.value and 'STD' in celda_binarizacion.value:
            for cell in worksheet[row_idx]: cell.fill = fill_std
        for col_idx in range(2, worksheet.max_column + 1):
            worksheet.cell(row=row_idx, column=col_idx).number_format = formato_numero

    columnas_fitness = {
        'B': 'Avg. Fitness Promedio',
        'C': 'Best Fitness Promedio',
        'D': 'Best Fitness Máximo'
    }
    for col_letter, col_name in columnas_fitness.items():
        mid_point = df_estadisticas[col_name].mean()
        # Se configura la regla con 3 puntos: start, mid, end
        regla_3_colores = ColorScaleRule(
            start_type='min', start_color='F8696B',
            mid_type='num', mid_value=mid_point, mid_color='FFFF00',
            end_type='max', end_color='63BE7B'
        )
        rango = f'{col_letter}{start_row_stats}:{col_letter}{end_row_stats}'
        worksheet.conditional_formatting.add(rango, regla_3_colores)

# --- FUNCIONES DE CÁLCULO DE RESUMEN ---
def crear_resumen_victorias_binarizacion(df_total):
    df_total['Instancia_Base'] = df_total['Instance'].str.split(' \(').str[0]
    idx_ganadores = df_total.groupby('Instancia_Base')['Best Fitness'].idxmax()
    df_ganadores = df_total.loc[idx_ganadores]
    df_ganadores['Binarizacion_Ganadora'] = df_ganadores['Instance'].str.extract(r'\((\S+)\)')
    resumen_victorias = df_ganadores['Binarizacion_Ganadora'].value_counts().reset_index()
    resumen_victorias.columns = ['Técnica de Binarización', 'Victorias Absolutas']
    return resumen_victorias.sort_values(by='Victorias Absolutas', ascending=False)

def crear_resumen_estadistico_binarizacion(df_total):
    df_total['Binarizacion'] = df_total['Instance'].str.extract(r'\((\S+)\)')
    agregaciones = {'Avg. Fitness': 'mean', 'Best Fitness': ['mean', 'max'], 'Avg. Time': 'mean'}
    resumen_stats = df_total.groupby('Binarizacion').agg(agregaciones).round(2)
    resumen_stats.columns = ['Avg. Fitness Promedio', 'Best Fitness Promedio', 'Best Fitness Máximo', 'Avg. Time Promedio']
    return resumen_stats.sort_values(by='Avg. Fitness Promedio', ascending=False)

def realizar_analisis_estadistico(df_global):
    df_global['Tipo'] = df_global['Binarizacion'].apply(lambda x: 'GVP' if 'GVP' in x else 'STD')
    gvp_data = df_global[df_global['Tipo'] == 'GVP']['Avg. Fitness']
    std_data = df_global[df_global['Tipo'] == 'STD']['Avg. Fitness']
    if gvp_data.empty or std_data.empty: return pd.DataFrame([{"Error": "No hay datos suficientes."}])
    stat, p_value = stats.mannwhitneyu(gvp_data, std_data, alternative='two-sided')
    conclusion = "Diferencia estadísticamente significativa" if p_value < 0.05 else "Sin diferencia estadísticamente significativa"
    resultado = {"Prueba": "Mann-Whitney U (GVP vs STD)", "p-valor": round(p_value, 5), "Conclusión (α=0.05)": conclusion}
    return pd.DataFrame([resultado])

# --- SECCIÓN DE GRÁFICOS ---
def crear_grafico_boxplot(df_global, ruta_imagen):
    plt.figure(figsize=(8, 5)); sns.set_style("whitegrid")
    if 'Tipo' not in df_global.columns:
        df_global['Tipo'] = df_global['Binarizacion'].apply(lambda x: 'GVP' if 'GVP' in x else 'STD')
    ax = sns.boxplot(x='Tipo', y='Avg. Fitness', data=df_global, palette=["#86C17E", "#7BA3D2"], hue='Tipo', legend=False)
    ax.set_title('Distribución de Fitness (GVP vs. STD)', fontsize=14, weight='bold')
    ax.set_xlabel('Tipo de Binarización', fontsize=10); ax.set_ylabel('Avg. Fitness Obtenido', fontsize=10)
    plt.tight_layout(); plt.savefig(ruta_imagen, dpi=200); plt.close()

def crear_grafico_barras_resumen(df_estadisticas, ruta_imagen):
    df_plot = df_estadisticas.reset_index()
    df_plot['Tipo'] = df_plot['Binarizacion'].apply(lambda x: 'GVP' if 'GVP' in x else 'STD')
    colores = df_plot['Tipo'].map({'GVP': '#86C17E', 'STD': '#7BA3D2'})
    plt.figure(figsize=(10, 6)); sns.set_style("whitegrid")
    ax = sns.barplot(y='Binarizacion', x='Avg. Fitness Promedio', data=df_plot, palette=colores.tolist(), hue='Binarizacion', legend=False)
    ax.set_title('Rendimiento Promedio por Técnica de Binarización', fontsize=16, weight='bold')
    ax.set_xlabel('Avg. Fitness Promedio', fontsize=12); ax.set_ylabel('Técnica de Binarización', fontsize=12)
    min_val = df_plot['Avg. Fitness Promedio'].min(); max_val = df_plot['Avg. Fitness Promedio'].max()
    range_val = max_val - min_val
    ax.set_xlim(left=min_val - range_val * 0.05, right=max_val + range_val * 0.05)
    plt.tight_layout(); plt.savefig(ruta_imagen, dpi=200); plt.close()

def crear_grafico_barras_detalle(df_detalle, binarizacion_titulo, ruta_imagen):
    print(f"    > Generando gráfico para {binarizacion_titulo}...")
    plt.figure(figsize=(8, 5)); sns.set_style("whitegrid")
    ax = sns.barplot(y='Metaheuristic', x='Avg. Fitness', data=df_detalle, color='#5DADE2')
    ax.set_title(f'Rendimiento para {binarizacion_titulo}', fontsize=14, weight='bold')
    ax.set_xlabel('Avg. Fitness', fontsize=10); ax.set_ylabel('Metaheurística', fontsize=10)
    min_val = df_detalle['Avg. Fitness'].min(); max_val = df_detalle['Avg. Fitness'].max()
    range_val = max_val - min_val
    ax.set_xlim(left=min_val - range_val * 0.1, right=max_val + range_val * 0.01)
    plt.tight_layout(); plt.savefig(ruta_imagen, dpi=200); plt.close()

# --- FUNCIÓN DE ESCRITURA DE RESUMEN ---
def escribir_hoja_resumen_global(writer, df_global, escala):
    sheet_name = "Resumen Global y Gráficos"
    print(f"\n[INFO] Creando la hoja principal '{sheet_name}'...")
    current_row = 0; df_victorias = crear_resumen_victorias_binarizacion(df_global)
    pd.DataFrame(["Victorias Absolutas por Técnica de Binarización"]).to_excel(writer, sheet_name=sheet_name, startrow=current_row, index=False, header=False)
    current_row += 2; df_victorias.to_excel(writer, sheet_name=sheet_name, startrow=current_row, index=False)
    current_row += len(df_victorias) + 3
    df_estadisticas = crear_resumen_estadistico_binarizacion(df_global)
    start_row_stats_data = current_row + 2
    pd.DataFrame(["Análisis Estadístico por Técnica de Binarización"]).to_excel(writer, sheet_name=sheet_name, startrow=current_row, index=False, header=False)
    current_row += 2; df_estadisticas.to_excel(writer, sheet_name=sheet_name, startrow=current_row)
    end_row_stats_data = current_row + len(df_estadisticas)
    current_row += len(df_estadisticas) + 3
    df_test = realizar_analisis_estadistico(df_global)
    pd.DataFrame(["Prueba de Hipótesis (GVP vs. STD)"]).to_excel(writer, sheet_name=sheet_name, startrow=current_row, index=False, header=False)
    current_row += 2; df_test.to_excel(writer, sheet_name=sheet_name, startrow=current_row, index=False)
    rutas_generadas = []; ruta_boxplot = 'resumen_boxplot.png'; rutas_generadas.append(ruta_boxplot)
    crear_grafico_boxplot(df_global, ruta_boxplot)
    worksheet = writer.sheets[sheet_name]; img_boxplot = Image(ruta_boxplot)
    img_boxplot.width, img_boxplot.height = (img_boxplot.width * escala, img_boxplot.height * escala)
    worksheet.add_image(img_boxplot, 'G2')
    ruta_barras = 'resumen_barras.png'; rutas_generadas.append(ruta_barras)
    crear_grafico_barras_resumen(df_estadisticas, ruta_barras)
    img_barras = Image(ruta_barras)
    img_barras.width, img_barras.height = (img_barras.width * escala, img_barras.height * escala)
    worksheet.add_image(img_barras, 'G22')
    aplicar_estilos_basicos(worksheet)
    aplicar_estilos_avanzados_resumen(worksheet, df_estadisticas, start_row_stats_data, end_row_stats_data)
    return rutas_generadas

# --- FUNCIÓN PRINCIPAL ---
def consolidar_resultados_agrupados():
    print("[FASE 1] Procesando todos los archivos de resultados...")
    datos_agrupados = {}; patron = re.compile(r"resumen_(fitness|times)_KP_(.+)_(.+)\.csv")
    for filename in os.listdir(RESUMEN_DIR):
        match = patron.match(filename)
        if match:
            tipo_resumen, instancia, binarizacion = match.groups()
            if instancia not in datos_agrupados: datos_agrupados[instancia] = {}
            if binarizacion not in datos_agrupados[instancia]: datos_agrupados[instancia][binarizacion] = {}
            try:
                df = pd.read_csv(os.path.join(RESUMEN_DIR, filename)); df.columns = df.columns.str.strip()
                datos_agrupados[instancia][binarizacion][tipo_resumen] = df
            except Exception as e: print(f"[ADVERTENCIA] No se pudo leer {filename}: {e}")
    if not datos_agrupados: print("[INFO] No se encontraron archivos."); return
    datos_procesados = {}
    for instancia, binarizaciones in datos_agrupados.items():
        datos_procesados[instancia] = {}
        for binarizacion, dataframes in binarizaciones.items():
            if 'fitness' not in dataframes or 'times' not in dataframes: continue
            df_fitness = dataframes['fitness']; df_times = dataframes['times']
            df_final = pd.merge(df_fitness, df_times, on='instance', how='outer')
            df_final.rename(columns={'instance': 'Metaheuristic', 'best': 'Best Fitness','avg. fitness': 'Avg. Fitness', 'std fitness': 'Std. Dev. Fitness','min time (s)': 'Min Time', 'avg. time (s)': 'Avg. Time','std time (s)': 'Std. Dev. Time'}, inplace=True)
            df_final['Population'] = PARAM_POBLACION; df_final['Iterations'] = PARAM_ITERACIONES
            df_final['Problem'] = PARAM_PROBLEMA; df_final['Instance'] = f"{instancia} ({binarizacion})"
            df_final['Binarizacion'] = binarizacion 
            orden_fitness_asc = False if PARAM_PROBLEMA == 'KP' else True
            df_final = df_final.sort_values(by=['Avg. Fitness', 'Avg. Time'], ascending=[orden_fitness_asc, True])
            df_final.reset_index(drop=True, inplace=True); df_final['Ranking'] = df_final.index + 1
            column_order = ['Metaheuristic', 'Population', 'Iterations', 'Problem', 'Instance', 'Binarizacion', 'Best Fitness', 'Avg. Fitness', 'Std. Dev. Fitness','Min Time', 'Avg. Time', 'Std. Dev. Time', 'Ranking']
            datos_procesados[instancia][binarizacion] = df_final[column_order]
    
    print("\n[FASE 2] Escribiendo archivo Excel final...")
    rutas_graficos_a_borrar = []
    escala_resumen = 0.25
    escala_detalle = 0.5
    
    try:
        with pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl') as writer:
            todos_los_resultados_dfs = [df for binarizaciones in datos_procesados.values() for df in binarizaciones.values()]
            if todos_los_resultados_dfs:
                df_global = pd.concat(todos_los_resultados_dfs, ignore_index=True)
                rutas_generadas = escribir_hoja_resumen_global(writer, df_global, escala=escala_resumen)
                rutas_graficos_a_borrar.extend(rutas_generadas)

            print("\n[INFO] Escribiendo hojas de detalles con sus gráficos...")
            for instancia, binarizaciones_proc in datos_procesados.items():
                nombre_hoja_detalles = instancia[:31]
                print(f"  > Creando hoja de detalles '{nombre_hoja_detalles}'...")
                current_row = 0
                for binarizacion, df_procesado in binarizaciones_proc.items():
                    pd.DataFrame([f"Resultados para Binarización: {binarizacion}"]).to_excel(writer, sheet_name=nombre_hoja_detalles, startrow=current_row, index=False, header=False)
                    current_row += 2
                    ruta_detalle = f"temp_plot_{instancia}_{binarizacion}.png"
                    rutas_graficos_a_borrar.append(ruta_detalle)
                    crear_grafico_barras_detalle(df_procesado, binarizacion, ruta_detalle)
                    df_procesado.to_excel(writer, sheet_name=nombre_hoja_detalles, startrow=current_row, index=False)
                    worksheet = writer.sheets[nombre_hoja_detalles]
                    img = Image(ruta_detalle)
                    img.width, img.height = (img.width * escala_detalle, img.height * escala_detalle)
                    worksheet.add_image(img, f'N{current_row + 1}')
                    current_row += len(df_procesado) + 2
                aplicar_estilos_basicos(writer.sheets[nombre_hoja_detalles])
        
        print("\n[ÉXITO] El reporte gráfico completo ha sido generado.")
    finally:
        print("\n[INFO] Limpiando archivos de imagen temporales...")
        for ruta in rutas_graficos_a_borrar:
            if os.path.exists(ruta):
                os.remove(ruta)
                print(f"  > Archivo '{ruta}' eliminado.")

if __name__ == "__main__":
    consolidar_resultados_agrupados()